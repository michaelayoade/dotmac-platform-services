"""
Data Export Module

Comprehensive exporters for CSV, JSON, Excel and other formats with streaming support,
compression, and progress tracking.
"""

import csv
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from io import StringIO, BytesIO
from pathlib import Path
from typing import Any, AsyncGenerator, BinaryIO, TextIO
import zipfile
import gzip
import bz2

import aiofiles
import openpyxl
import yaml
from pydantic import BaseModel, ConfigDict, Field, field_validator

from .base import (
    BaseExporter,
    CompressionType,
    DataBatch,
    DataFormat,
    DataRecord,
    ExportError,
    FormatError,
    ProgressCallback,
    ProgressInfo,
    StreamingError,
    TransferConfig,
    TransferStatus,
    create_operation_id,
)
from .progress import ProgressTracker


async def _write_to_stream(stream: Any, content: str) -> None:
    """Write content to stream, handling both sync and async streams."""
    if hasattr(stream, "write"):
        write_result = stream.write(content)
        if hasattr(write_result, '__await__'):
            await write_result
    else:
        stream.write(content)


class ExportOptions(BaseModel):
    """Export operation options."""

    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    # CSV specific options
    delimiter: str = Field(",", description="CSV delimiter")
    quotechar: str = Field('"', description="CSV quote character")
    quoting: int = Field(csv.QUOTE_MINIMAL, description="CSV quoting behavior")
    line_terminator: str = Field("\n", description="Line terminator")
    include_headers: bool = Field(True, description="Include header row")

    # JSON specific options
    json_indent: int | None = Field(2, description="JSON indentation")
    json_ensure_ascii: bool = Field(False, description="Ensure ASCII encoding")
    json_sort_keys: bool = Field(False, description="Sort JSON keys")
    json_lines: bool = Field(False, description="Output as JSON Lines format")

    # Excel specific options
    sheet_name: str = Field("Sheet1", description="Excel sheet name")
    freeze_panes: str | None = Field(None, description="Freeze panes (e.g., 'A2')")
    auto_filter: bool = Field(True, description="Enable auto filter")

    # XML specific options
    xml_root_element: str = Field("root", description="Root element name")
    xml_record_element: str = Field("record", description="Record element name")
    xml_pretty_print: bool = Field(True, description="Pretty print XML")

    # General options
    date_format: str = Field("%Y-%m-%d %H:%M:%S", description="Date format string")
    null_value: str = Field("", description="Representation for null values")
    include_metadata: bool = Field(False, description="Include record metadata")
    max_file_size: int = Field(100 * 1024 * 1024, description="Maximum file size in bytes")

    @field_validator("delimiter")
    @classmethod
    def validate_delimiter(cls, v: str) -> str:
        if len(v) != 1:
            raise ValueError("Delimiter must be a single character")
        return v


class CSVExporter(BaseExporter):
    """CSV data exporter with streaming support."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ExportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ExportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process method - for exporters, this typically passes through data batches."""
        if hasattr(source, '__aiter__'):
            # Source is an async generator of data batches
            async for batch in source:
                yield batch
        else:
            # Source is a single data batch or list of records
            if isinstance(source, DataBatch):
                yield source
            else:
                # Convert records to a single batch
                records = source if isinstance(source, list) else [source]
                batch = DataBatch(
                    batch_number=1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records)
                )
                yield batch

    async def export_to_file(
        self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs
    ) -> ProgressInfo:
        """Export data to CSV file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Apply compression if needed
            if self.config.compression != CompressionType.NONE:
                file_path = await self._get_compressed_path(file_path)

            async with aiofiles.open(file_path, "w", encoding=self.config.encoding) as f:
                await self.export_to_stream(data, f, **kwargs)

            self._progress.status = TransferStatus.COMPLETED
            return self._progress

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ExportError(f"Failed to export to CSV file {file_path}: {e}") from e

    async def export_to_stream(
        self, data: AsyncGenerator[DataBatch, None], stream: TextIO, **kwargs
    ) -> ProgressInfo:
        """Export data to CSV stream."""
        try:
            csv_buffer = StringIO()
            csv_writer = None
            headers_written = False
            all_fieldnames = set()
            fieldnames_order = []  # To preserve order of first appearance
            batches_to_process = []

            # First pass: collect all unique field names and store batches
            async for batch in data:
                if self._cancelled:
                    break
                batches_to_process.append(batch)

                for record in batch.records:
                    if not record.is_valid and self.config.skip_invalid:
                        continue

                    # Collect field names preserving order of first appearance
                    for field_name in record.data.keys():
                        if field_name not in all_fieldnames:
                            all_fieldnames.add(field_name)
                            fieldnames_order.append(field_name)

                    # Add metadata fields
                    if self.options.include_metadata and record.metadata:
                        for meta_key in record.metadata.keys():
                            meta_field = f"meta_{meta_key}"
                            if meta_field not in all_fieldnames:
                                all_fieldnames.add(meta_field)
                                fieldnames_order.append(meta_field)

            # Use order-preserving fieldnames list
            fieldnames = fieldnames_order

            # Initialize CSV writer with complete fieldnames
            if fieldnames:
                csv_writer = csv.DictWriter(
                    csv_buffer,
                    fieldnames=fieldnames,
                    delimiter=self.options.delimiter,
                    quotechar=self.options.quotechar,
                    quoting=self.options.quoting,
                    lineterminator=self.options.line_terminator,
                    extrasaction='ignore',  # Ignore extra fields
                    restval=''  # Fill missing fields with empty string
                )

                if self.options.include_headers:
                    csv_writer.writeheader()
                    headers_written = True

            # Second pass: write all data
            for batch in batches_to_process:
                if self._cancelled:
                    break

                for record in batch.records:
                    if not record.is_valid and self.config.skip_invalid:
                        continue

                    # Prepare row data with all fieldnames
                    row_data = {}
                    for fieldname in fieldnames:
                        if fieldname.startswith("meta_"):
                            # Handle metadata fields
                            meta_key = fieldname[5:]  # Remove "meta_" prefix
                            if self.options.include_metadata and record.metadata and meta_key in record.metadata:
                                row_data[fieldname] = self._format_value(record.metadata[meta_key])
                            else:
                                row_data[fieldname] = ''
                        else:
                            # Handle regular data fields
                            if fieldname in record.data:
                                row_data[fieldname] = self._format_value(record.data[fieldname])
                            else:
                                row_data[fieldname] = ''

                    # Write row
                    if csv_writer:
                        csv_writer.writerow(row_data)

                # Write buffer to stream
                csv_content = csv_buffer.getvalue()
                await _write_to_stream(stream, csv_content)

                csv_buffer.seek(0)
                csv_buffer.truncate(0)

                # Update progress
                self._update_progress(
                    processed_records=self._progress.processed_records + len(batch.records),
                    current_batch=self._progress.current_batch + 1,
                    bytes_processed=self._progress.bytes_processed + batch.total_size,
                )

            return self._progress

        except Exception as e:
            raise ExportError(f"Failed to export to CSV stream: {e}") from e

    def _format_value(self, value: Any) -> str:
        """Format value for CSV output."""
        if value is None:
            return self.options.null_value

        if isinstance(value, datetime):
            return value.strftime(self.options.date_format)

        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)

        return str(value)


class JSONExporter(BaseExporter):
    """JSON data exporter with streaming support."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ExportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ExportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process method - for exporters, this typically passes through data batches."""
        if hasattr(source, '__aiter__'):
            # Source is an async generator of data batches
            async for batch in source:
                yield batch
        else:
            # Source is a single data batch or list of records
            if isinstance(source, DataBatch):
                yield source
            else:
                # Convert records to a single batch
                records = source if isinstance(source, list) else [source]
                batch = DataBatch(
                    batch_number=1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records)
                )
                yield batch

    async def export_to_file(
        self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs
    ) -> ProgressInfo:
        """Export data to JSON file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Apply compression if needed
            if self.config.compression != CompressionType.NONE:
                file_path = await self._get_compressed_path(file_path)

            async with aiofiles.open(file_path, "w", encoding=self.config.encoding) as f:
                await self.export_to_stream(data, f, **kwargs)

            self._progress.status = TransferStatus.COMPLETED
            return self._progress

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ExportError(f"Failed to export to JSON file {file_path}: {e}") from e

    async def export_to_stream(
        self, data: AsyncGenerator[DataBatch, None], stream: TextIO, **kwargs
    ) -> ProgressInfo:
        """Export data to JSON stream."""
        try:
            if self.options.json_lines:
                await self._export_jsonl(data, stream)
            else:
                await self._export_json(data, stream)

            return self._progress

        except Exception as e:
            raise ExportError(f"Failed to export to JSON stream: {e}") from e

    async def _export_json(self, data: AsyncGenerator[DataBatch, None], stream: TextIO) -> None:
        """Export as standard JSON array."""
        # Write opening bracket
        if hasattr(stream, "write"):
            await _write_to_stream(stream, "[\n" if self.options.json_indent else "[")
        else:
            stream.write("[\n" if self.options.json_indent else "[")

        first_record = True

        async for batch in data:
            if self._cancelled:
                break

            for record in batch.records:
                if not record.is_valid and self.config.skip_invalid:
                    continue

                # Prepare record data
                record_data = dict(record.data)
                if self.options.include_metadata and record.metadata:
                    record_data["_metadata"] = record.metadata

                # Convert to JSON
                record_json = json.dumps(
                    record_data,
                    indent=self.options.json_indent,
                    ensure_ascii=self.options.json_ensure_ascii,
                    sort_keys=self.options.json_sort_keys,
                    default=self._json_serializer,
                )

                # Add comma separator
                if not first_record:
                    content = ",\n" + record_json if self.options.json_indent else "," + record_json
                else:
                    content = record_json
                    first_record = False

                # Write to stream
                if hasattr(stream, "write"):
                    await _write_to_stream(stream, content)
                else:
                    stream.write(content)

            # Update progress
            self._update_progress(
                processed_records=self._progress.processed_records + len(batch.records),
                current_batch=self._progress.current_batch + 1,
                bytes_processed=self._progress.bytes_processed + batch.total_size,
            )

        # Write closing bracket
        if hasattr(stream, "write"):
            await _write_to_stream(stream, "\n]" if self.options.json_indent else "]")
        else:
            stream.write("\n]" if self.options.json_indent else "]")

    async def _export_jsonl(self, data: AsyncGenerator[DataBatch, None], stream: TextIO) -> None:
        """Export as JSON Lines format."""
        async for batch in data:
            if self._cancelled:
                break

            for record in batch.records:
                if not record.is_valid and self.config.skip_invalid:
                    continue

                # Prepare record data
                record_data = dict(record.data)
                if self.options.include_metadata and record.metadata:
                    record_data["_metadata"] = record.metadata

                # Convert to JSON and write line
                record_json = json.dumps(
                    record_data,
                    ensure_ascii=self.options.json_ensure_ascii,
                    sort_keys=self.options.json_sort_keys,
                    default=self._json_serializer,
                )

                line = record_json + "\n"
                if hasattr(stream, "write"):
                    await _write_to_stream(stream, line)
                else:
                    stream.write(line)

            # Update progress
            self._update_progress(
                processed_records=self._progress.processed_records + len(batch.records),
                current_batch=self._progress.current_batch + 1,
                bytes_processed=self._progress.bytes_processed + batch.total_size,
            )

    def _json_serializer(self, obj: Any) -> Any:
        """Custom JSON serializer for complex objects."""
        if isinstance(obj, datetime):
            return obj.strftime(self.options.date_format)

        if hasattr(obj, "model_dump"):  # Pydantic models
            return obj.model_dump()

        if hasattr(obj, "__dict__"):
            return obj.__dict__

        return str(obj)


class ExcelExporter(BaseExporter):
    """Excel data exporter."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ExportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ExportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process method - for exporters, this typically passes through data batches."""
        if hasattr(source, '__aiter__'):
            # Source is an async generator of data batches
            async for batch in source:
                yield batch
        else:
            # Source is a single data batch or list of records
            if isinstance(source, DataBatch):
                yield source
            else:
                # Convert records to a single batch
                records = source if isinstance(source, list) else [source]
                batch = DataBatch(
                    batch_number=1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records)
                )
                yield batch

    async def export_to_file(
        self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs
    ) -> ProgressInfo:
        """Export data to Excel file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = self.options.sheet_name

            await self._populate_worksheet(worksheet, data)

            # Apply formatting
            if self.options.freeze_panes:
                worksheet.freeze_panes = self.options.freeze_panes

            # Apply auto filter if enabled and we have actual data
            if self.options.auto_filter:
                try:
                    # Check if max_row is a real integer value greater than 1
                    if isinstance(worksheet.max_row, int) and worksheet.max_row > 1:
                        worksheet.auto_filter.ref = worksheet.dimensions
                except (TypeError, AttributeError):
                    # Skip auto filter if worksheet.max_row is not a valid value
                    pass

            # Save workbook
            workbook.save(file_path)

            self._progress.status = TransferStatus.COMPLETED
            return self._progress

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ExportError(f"Failed to export to Excel file {file_path}: {e}") from e

    async def export_to_stream(
        self, data: AsyncGenerator[DataBatch, None], stream: BinaryIO, **kwargs
    ) -> ProgressInfo:
        """Export data to Excel stream."""
        # Excel export to stream is not reliable due to openpyxl limitations
        raise ExportError("Excel export to stream not supported")

    async def _populate_worksheet(self, worksheet, data: AsyncGenerator[DataBatch, None]) -> None:
        """Populate worksheet with data."""
        current_row = 1
        headers_written = False

        async for batch in data:
            if self._cancelled:
                break

            for record in batch.records:
                if not record.is_valid and self.config.skip_invalid:
                    continue

                # Write headers
                if not headers_written:
                    headers = list(record.data.keys())
                    if self.options.include_metadata and record.metadata:
                        headers.extend([f"meta_{k}" for k in record.metadata.keys()])

                    for col, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.value = header
                        cell.font = openpyxl.styles.Font(bold=True)

                    current_row += 1
                    headers_written = True

                # Write data row
                col = 1
                for key, value in record.data.items():
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.value = self._format_excel_value(value)
                    col += 1

                # Add metadata if requested
                if self.options.include_metadata and record.metadata:
                    for key, value in record.metadata.items():
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.value = self._format_excel_value(value)
                        col += 1

                current_row += 1

            # Update progress
            self._update_progress(
                processed_records=self._progress.processed_records + len(batch.records),
                current_batch=self._progress.current_batch + 1,
                bytes_processed=self._progress.bytes_processed + batch.total_size,
            )

    def _format_excel_value(self, value: Any) -> Any:
        """Format value for Excel output."""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)

        return value


class XMLExporter(BaseExporter):
    """XML data exporter."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ExportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ExportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process method - for exporters, this typically passes through data batches."""
        if hasattr(source, '__aiter__'):
            # Source is an async generator of data batches
            async for batch in source:
                yield batch
        else:
            # Source is a single data batch or list of records
            if isinstance(source, DataBatch):
                yield source
            else:
                # Convert records to a single batch
                records = source if isinstance(source, list) else [source]
                batch = DataBatch(
                    batch_number=1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records)
                )
                yield batch

    async def export_to_file(
        self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs
    ) -> ProgressInfo:
        """Export data to XML file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Apply compression if needed
            if self.config.compression != CompressionType.NONE:
                file_path = await self._get_compressed_path(file_path)

            async with aiofiles.open(file_path, "w", encoding=self.config.encoding) as f:
                await self.export_to_stream(data, f, **kwargs)

            self._progress.status = TransferStatus.COMPLETED
            return self._progress

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ExportError(f"Failed to export to XML file {file_path}: {e}") from e

    async def export_to_stream(
        self, data: AsyncGenerator[DataBatch, None], stream: TextIO, **kwargs
    ) -> ProgressInfo:
        """Export data to XML stream."""
        try:
            # Create root element
            root = ET.Element(self.options.xml_root_element)

            async for batch in data:
                if self._cancelled:
                    break

                for record in batch.records:
                    if not record.is_valid and self.config.skip_invalid:
                        continue

                    # Create record element
                    record_element = ET.SubElement(root, self.options.xml_record_element)
                    self._populate_element_from_dict(record.data, record_element)

                    # Add metadata if requested
                    if self.options.include_metadata and record.metadata:
                        metadata_element = ET.SubElement(record_element, "metadata")
                        self._populate_element_from_dict(record.metadata, metadata_element)

                # Update progress
                self._update_progress(
                    processed_records=self._progress.processed_records + len(batch.records),
                    current_batch=self._progress.current_batch + 1,
                    bytes_processed=self._progress.bytes_processed + batch.total_size,
                )

            # Write XML to stream
            if self.options.xml_pretty_print:
                self._indent_xml(root)

            xml_string = ET.tostring(root, encoding="unicode")
            xml_content = f'<?xml version="1.0" encoding="{self.config.encoding}"?>\n{xml_string}'

            if hasattr(stream, "write"):
                await _write_to_stream(stream, xml_content)
            else:
                stream.write(xml_content)

            return self._progress

        except Exception as e:
            raise ExportError(f"Failed to export to XML stream: {e}") from e

    def _dict_to_element(self, element_name: str, data: dict[str, Any]) -> ET.Element:
        """Convert dictionary to XML element."""
        # Clean element name for XML
        clean_name = str(element_name).replace(" ", "_").replace("-", "_")
        root_element = ET.Element(clean_name)

        for key, value in data.items():
            # Handle attributes (keys starting with @)
            if key.startswith("@"):
                attr_name = key[1:]  # Remove @ prefix
                root_element.set(attr_name, str(value))
                continue

            # Clean key name for XML
            clean_key = str(key).replace(" ", "_").replace("-", "_")

            if isinstance(value, dict):
                sub_element = ET.SubElement(root_element, clean_key)
                # Recursively add dictionary content
                for sub_key, sub_value in value.items():
                    if sub_key.startswith("@"):
                        sub_element.set(sub_key[1:], str(sub_value))
                    else:
                        sub_sub_element = ET.SubElement(sub_element, str(sub_key).replace(" ", "_").replace("-", "_"))
                        sub_sub_element.text = self._format_xml_value(sub_value)
            elif isinstance(value, list):
                for item in value:
                    element = ET.SubElement(root_element, clean_key)
                    if isinstance(item, dict):
                        # For dictionary items in list, add their content directly to the element
                        for item_key, item_value in item.items():
                            if item_key.startswith("@"):
                                element.set(item_key[1:], str(item_value))
                            else:
                                sub_elem = ET.SubElement(element, str(item_key).replace(" ", "_").replace("-", "_"))
                                sub_elem.text = self._format_xml_value(item_value)
                    else:
                        element.text = self._format_xml_value(item)
            else:
                element = ET.SubElement(root_element, clean_key)
                element.text = self._format_xml_value(value)

        return root_element

    def _populate_element_from_dict(self, data: dict[str, Any], parent: ET.Element) -> None:
        """Populate an existing XML element with dictionary data."""
        for key, value in data.items():
            # Handle attributes (keys starting with @)
            if key.startswith("@"):
                attr_name = key[1:]  # Remove @ prefix
                parent.set(attr_name, str(value))
                continue

            # Clean key name for XML
            clean_key = str(key).replace(" ", "_").replace("-", "_")

            if isinstance(value, dict):
                sub_element = ET.SubElement(parent, clean_key)
                self._populate_element_from_dict(value, sub_element)
            elif isinstance(value, list):
                for item in value:
                    element = ET.SubElement(parent, clean_key)
                    if isinstance(item, dict):
                        self._populate_element_from_dict(item, element)
                    else:
                        element.text = self._format_xml_value(item)
            else:
                element = ET.SubElement(parent, clean_key)
                element.text = self._format_xml_value(value)

    def _format_xml_value(self, value: Any) -> str:
        """Format value for XML output."""
        if value is None:
            return self.options.null_value

        if isinstance(value, datetime):
            return value.strftime(self.options.date_format)

        return str(value)

    def _indent_xml(self, elem: ET.Element, level: int = 0) -> None:
        """Add pretty-print indentation to XML."""
        indent = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = indent + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
            for elem in elem:
                self._indent_xml(elem, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = indent


class YAMLExporter(BaseExporter):
    """YAML data exporter."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ExportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ExportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process method - for exporters, this typically passes through data batches."""
        if hasattr(source, '__aiter__'):
            # Source is an async generator of data batches
            async for batch in source:
                yield batch
        else:
            # Source is a single data batch or list of records
            if isinstance(source, DataBatch):
                yield source
            else:
                # Convert records to a single batch
                records = source if isinstance(source, list) else [source]
                batch = DataBatch(
                    batch_number=1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records)
                )
                yield batch

    async def export_to_file(
        self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs
    ) -> ProgressInfo:
        """Export data to YAML file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Collect all data
            all_records = []
            async for batch in data:
                if self._cancelled:
                    break

                for record in batch.records:
                    if not record.is_valid and self.config.skip_invalid:
                        continue

                    record_data = dict(record.data)
                    if self.options.include_metadata and record.metadata:
                        record_data["_metadata"] = record.metadata

                    all_records.append(record_data)

                # Update progress
                self._update_progress(
                    processed_records=self._progress.processed_records + len(batch.records),
                    current_batch=self._progress.current_batch + 1,
                    bytes_processed=self._progress.bytes_processed + batch.total_size,
                )

            # Write YAML
            async with aiofiles.open(file_path, "w", encoding=self.config.encoding) as f:
                yaml_content = yaml.dump(
                    all_records,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=self.options.json_sort_keys,
                )
                await _write_to_stream(f, yaml_content)

            self._progress.status = TransferStatus.COMPLETED
            return self._progress

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ExportError(f"Failed to export to YAML file {file_path}: {e}") from e

    async def export_to_stream(
        self, data: AsyncGenerator[DataBatch, None], stream: TextIO, **kwargs
    ) -> ProgressInfo:
        """Export data to YAML stream."""
        try:
            # Collect all data
            all_records = []
            async for batch in data:
                if self._cancelled:
                    break

                for record in batch.records:
                    if not record.is_valid and self.config.skip_invalid:
                        continue

                    record_data = dict(record.data)
                    if self.options.include_metadata and record.metadata:
                        record_data["_metadata"] = record.metadata

                    all_records.append(record_data)

                # Update progress
                self._update_progress(
                    processed_records=self._progress.processed_records + len(batch.records),
                    current_batch=self._progress.current_batch + 1,
                    bytes_processed=self._progress.bytes_processed + batch.total_size,
                )

            # Write YAML to stream
            yaml_content = yaml.dump(
                all_records,
                default_flow_style=False,
                allow_unicode=True,
                sort_keys=self.options.json_sort_keys,
            )

            if hasattr(stream, "write"):
                await _write_to_stream(stream, yaml_content)
            else:
                stream.write(yaml_content)

            return self._progress

        except Exception as e:
            raise ExportError(f"Failed to export to YAML stream: {e}") from e


# Compression utilities


async def compress_file(input_path: Path, output_path: Path, compression: CompressionType) -> None:
    """Compress a file using the specified compression type."""
    if compression == CompressionType.GZIP:
        async with aiofiles.open(input_path, "rb") as f_in:
            with gzip.open(output_path, "wb") as f_out:
                content = await f_in.read()
                f_out.write(content)

    elif compression == CompressionType.BZIP2:
        async with aiofiles.open(input_path, "rb") as f_in:
            with bz2.open(output_path, "wb") as f_out:
                content = await f_in.read()
                f_out.write(content)

    elif compression == CompressionType.ZIP:
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.write(input_path, input_path.name)


# Factory functions


def create_exporter(
    data_format: DataFormat,
    config: TransferConfig | None = None,
    options: ExportOptions | None = None,
    progress_callback: ProgressCallback | None = None,
) -> BaseExporter:
    """Create an appropriate exporter for the given format."""
    exporters = {
        DataFormat.CSV: CSVExporter,
        DataFormat.JSON: JSONExporter,
        DataFormat.JSONL: lambda *args, **kwargs: JSONExporter(
            *args,
            options=ExportOptions(**(options.model_dump() if options else {}), json_lines=True),
            **kwargs,
        ),
        DataFormat.EXCEL: ExcelExporter,
        DataFormat.XML: XMLExporter,
        DataFormat.YAML: YAMLExporter,
    }

    if data_format not in exporters:
        raise FormatError(f"Unsupported format: {data_format}")

    return exporters[data_format](config, options, progress_callback)


async def export_data(
    data: AsyncGenerator[DataBatch, None],
    file_path: Path,
    data_format: DataFormat,
    config: TransferConfig | None = None,
    options: ExportOptions | None = None,
    progress_callback: ProgressCallback | None = None,
) -> ProgressInfo:
    """Export data to file with the specified format."""
    exporter = create_exporter(data_format, config, options, progress_callback)
    return await exporter.export_to_file(data, file_path)


# Mixin for compression support


class CompressionMixin:
    """Mixin to add compression support to exporters."""

    async def _get_compressed_path(self, file_path: Path) -> Path:
        """Get compressed file path based on compression type."""
        if self.config.compression == CompressionType.GZIP:
            return file_path.with_suffix(file_path.suffix + ".gz")
        elif self.config.compression == CompressionType.BZIP2:
            return file_path.with_suffix(file_path.suffix + ".bz2")
        elif self.config.compression == CompressionType.ZIP:
            return file_path.with_suffix(".zip")
        else:
            return file_path

    async def _compress_if_needed(self, file_path: Path) -> Path:
        """Compress file if compression is enabled."""
        if self.config.compression == CompressionType.NONE:
            return file_path

        compressed_path = await self._get_compressed_path(file_path)
        await compress_file(file_path, compressed_path, self.config.compression)

        # Remove original file
        file_path.unlink()

        return compressed_path


# Add compression mixin to base exporters
for exporter_class in [CSVExporter, JSONExporter, XMLExporter, YAMLExporter]:
    # Add mixin methods to exporter classes
    exporter_class._get_compressed_path = CompressionMixin._get_compressed_path
    exporter_class._compress_if_needed = CompressionMixin._compress_if_needed
