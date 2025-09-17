"""
Data Import Module

Comprehensive importers for CSV, JSON, Excel and other formats with streaming support,
data validation, and error recovery.
"""

import csv
import json
import xml.etree.ElementTree as ET
from io import StringIO, TextIOWrapper
from pathlib import Path
from typing import Any, AsyncGenerator, BinaryIO, TextIO

import aiofiles
import openpyxl
import pandas as pd
from pydantic import BaseModel, ConfigDict, Field, field_validator

from .base import (
    BaseImporter,
    CompressionType,
    DataBatch,
    DataFormat,
    DataRecord,
    FormatError,
    ImportError,
    ProgressCallback,
    StreamingError,
    TransferConfig,
    TransferStatus,
    create_batches,
    create_operation_id,
)
from .progress import ProgressTracker


class ImportOptions(BaseModel):
    """Import operation options."""

    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    # CSV specific options
    delimiter: str = Field(",", description="CSV delimiter")
    quotechar: str = Field('"', description="CSV quote character")
    escapechar: str | None = Field(None, description="CSV escape character")
    skip_blank_lines: bool = Field(True, description="Skip blank lines")
    header_row: int = Field(0, description="Header row index (0-based)")

    # Excel specific options
    sheet_name: str | int | None = Field(None, description="Excel sheet name or index")
    skip_rows: int = Field(0, description="Number of rows to skip")

    # JSON specific options
    json_lines: bool = Field(False, description="Treat as JSON Lines format")
    json_path: str | None = Field(None, description="JSONPath to extract data")

    # XML specific options
    xml_root_element: str | None = Field(None, description="Root element for XML parsing")
    xml_record_element: str | None = Field(None, description="Record element for XML parsing")

    # General options
    sample_size: int = Field(1000, ge=1, description="Sample size for format detection")
    type_inference: bool = Field(True, description="Enable automatic type inference")
    date_formats: list[str] = Field(
        default_factory=lambda: ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"],
        description="Date format patterns to try",
    )

    @field_validator("delimiter")
    @classmethod
    def validate_delimiter(cls, v: str) -> str:
        if len(v) != 1:
            raise ValueError("Delimiter must be a single character")
        return v


class CSVImporter(BaseImporter):
    """CSV data importer with streaming support."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ImportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ImportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process import source and yield batches."""
        if isinstance(source, (str, Path)):
            # Source is a file path
            async for batch in self.import_from_file(Path(source), **kwargs):
                yield batch
        else:
            # Source is a stream
            async for batch in self.import_from_stream(source, **kwargs):
                yield batch

    async def import_from_file(self, file_path: Path, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from CSV file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            async with aiofiles.open(file_path, "r", encoding=self.config.encoding) as f:
                async for batch in self.import_from_stream(f, **kwargs):
                    if self._cancelled:
                        break
                    yield batch

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ImportError(f"Failed to import CSV file {file_path}: {e}") from e

    async def import_from_stream(self, stream: TextIO, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from CSV stream."""
        try:
            # Read content
            if hasattr(stream, "read"):
                # Check if it's an async stream by trying to call read() and checking if it's a coroutine
                try:
                    read_result = stream.read()
                    # If it's a coroutine, await it
                    if hasattr(read_result, "__await__"):
                        content = await read_result
                    else:
                        content = read_result
                except Exception:
                    # Fallback to sync read
                    content = stream.read()
            else:
                content = stream.read()

            # Parse CSV
            csv_reader = csv.DictReader(
                StringIO(content),
                delimiter=self.options.delimiter,
                quotechar=self.options.quotechar,
                escapechar=self.options.escapechar,
            )

            # Skip to header row
            for _ in range(self.options.header_row):
                next(csv_reader, None)

            # Process records
            async for batch in self._process_csv_records(csv_reader):
                yield batch

        except Exception as e:
            raise ImportError(f"Failed to import from CSV stream: {e}") from e

    async def _process_csv_records(self, csv_reader) -> AsyncGenerator[DataBatch, None]:
        """Process CSV records into batches."""
        records = []
        row_number = self.options.header_row + 1

        for row_data in csv_reader:
            if self._cancelled:
                break

            # Skip blank lines
            if self.options.skip_blank_lines and not any(row_data.values()):
                continue

            # Create data record
            record = DataRecord(row_number=row_number, data=dict(row_data))

            # Apply type inference
            if self.options.type_inference:
                record.data = self._infer_types(record.data)

            # Validate and transform
            self._validate_record(record)
            record = self._transform_record(record)

            records.append(record)
            row_number += 1

            # Create batch when size limit reached
            if len(records) >= self.config.batch_size:
                batch = DataBatch(
                    batch_number=self._progress.current_batch + 1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records),
                )

                self._update_progress(
                    processed_records=self._progress.processed_records + len(records),
                    current_batch=self._progress.current_batch + 1,
                )

                yield batch
                records = []

        # Yield remaining records
        if records:
            batch = DataBatch(
                batch_number=self._progress.current_batch + 1,
                records=records,
                total_size=sum(len(str(r.data)) for r in records),
            )

            self._update_progress(
                processed_records=self._progress.processed_records + len(records),
                current_batch=self._progress.current_batch + 1,
            )

            yield batch

    def _infer_types(self, data: dict[str, Any]) -> dict[str, Any]:
        """Infer and convert data types."""
        converted = {}

        for key, value in data.items():
            if value is None or value == "":
                converted[key] = None
                continue

            str_value = str(value).strip()

            # Try integer
            try:
                if "." not in str_value:
                    converted[key] = int(str_value)
                    continue
            except (ValueError, TypeError):
                pass

            # Try float
            try:
                converted[key] = float(str_value)
                continue
            except (ValueError, TypeError):
                pass

            # Try boolean
            if str_value.lower() in ("true", "false", "yes", "no", "1", "0"):
                converted[key] = str_value.lower() in ("true", "yes", "1")
                continue

            # Try dates
            for date_format in self.options.date_formats:
                try:
                    from datetime import datetime

                    converted[key] = datetime.strptime(str_value, date_format)
                    break
                except ValueError:
                    continue
            else:
                # Keep as string
                converted[key] = str_value

        return converted


class JSONImporter(BaseImporter):
    """JSON data importer with streaming support."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ImportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ImportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process import source and yield batches."""
        if isinstance(source, (str, Path)):
            # Source is a file path
            async for batch in self.import_from_file(Path(source), **kwargs):
                yield batch
        else:
            # Source is a stream
            async for batch in self.import_from_stream(source, **kwargs):
                yield batch

    async def import_from_file(self, file_path: Path, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from JSON file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            async with aiofiles.open(file_path, "r", encoding=self.config.encoding) as f:
                async for batch in self.import_from_stream(f, **kwargs):
                    if self._cancelled:
                        break
                    yield batch

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ImportError(f"Failed to import JSON file {file_path}: {e}") from e

    async def import_from_stream(self, stream: TextIO, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from JSON stream."""
        try:
            # Read content
            if hasattr(stream, "read"):
                # Check if it's an async stream by trying to call read() and checking if it's a coroutine
                try:
                    read_result = stream.read()
                    # If it's a coroutine, await it
                    if hasattr(read_result, "__await__"):
                        content = await read_result
                    else:
                        content = read_result
                except Exception:
                    # Fallback to sync read
                    content = stream.read()
            else:
                content = stream.read()

            if self.options.json_lines:
                # Process JSON Lines format
                async for batch in self._process_jsonl(content):
                    yield batch
            else:
                # Process standard JSON
                async for batch in self._process_json(content):
                    yield batch

        except Exception as e:
            raise ImportError(f"Failed to import from JSON stream: {e}") from e

    async def _process_json(self, content: str) -> AsyncGenerator[DataBatch, None]:
        """Process standard JSON data."""
        try:
            data = json.loads(content)

            # Extract data using JSONPath if specified
            if self.options.json_path:
                data = self._extract_json_path(data, self.options.json_path)

            # Ensure data is iterable
            if not isinstance(data, list):
                data = [data]

            # Process records
            records = []
            for i, item in enumerate(data):
                if self._cancelled:
                    break

                record = DataRecord(
                    row_number=i + 1, data=item if isinstance(item, dict) else {"value": item}
                )

                self._validate_record(record)
                record = self._transform_record(record)
                records.append(record)

                # Create batch when size limit reached
                if len(records) >= self.config.batch_size:
                    batch = DataBatch(
                        batch_number=self._progress.current_batch + 1,
                        records=records,
                        total_size=len(json.dumps([r.data for r in records])),
                    )

                    self._update_progress(
                        processed_records=self._progress.processed_records + len(records),
                        current_batch=self._progress.current_batch + 1,
                    )

                    yield batch
                    records = []

            # Yield remaining records
            if records:
                batch = DataBatch(
                    batch_number=self._progress.current_batch + 1,
                    records=records,
                    total_size=len(json.dumps([r.data for r in records])),
                )

                self._update_progress(
                    processed_records=self._progress.processed_records + len(records),
                    current_batch=self._progress.current_batch + 1,
                )

                yield batch

        except json.JSONDecodeError as e:
            raise ImportError(f"Invalid JSON format: {e}") from e

    async def _process_jsonl(self, content: str) -> AsyncGenerator[DataBatch, None]:
        """Process JSON Lines format."""
        lines = content.strip().split("\n")
        records = []

        for i, line in enumerate(lines):
            if self._cancelled:
                break

            line = line.strip()
            if not line:
                continue

            try:
                data = json.loads(line)
                record = DataRecord(
                    row_number=i + 1, data=data if isinstance(data, dict) else {"value": data}
                )

                self._validate_record(record)
                record = self._transform_record(record)
                records.append(record)

                # Create batch when size limit reached
                if len(records) >= self.config.batch_size:
                    batch = DataBatch(
                        batch_number=self._progress.current_batch + 1,
                        records=records,
                        total_size=sum(len(json.dumps(r.data)) for r in records),
                    )

                    self._update_progress(
                        processed_records=self._progress.processed_records + len(records),
                        current_batch=self._progress.current_batch + 1,
                    )

                    yield batch
                    records = []

            except json.JSONDecodeError as e:
                # Create error record
                record = DataRecord(
                    row_number=i + 1,
                    data={"raw_line": line},
                    validation_errors=[f"JSON decode error: {e}"],
                    is_valid=False,
                )
                records.append(record)

        # Yield remaining records
        if records:
            batch = DataBatch(
                batch_number=self._progress.current_batch + 1,
                records=records,
                total_size=sum(len(json.dumps(r.data)) for r in records),
            )

            self._update_progress(
                processed_records=self._progress.processed_records + len(records),
                current_batch=self._progress.current_batch + 1,
            )

            yield batch

    def _extract_json_path(self, data: Any, json_path: str) -> Any:
        """Extract data using JSONPath (simplified implementation)."""
        # This is a basic implementation - for production use a proper JSONPath library
        path_parts = json_path.strip("$.").split(".")

        current = data
        for part in path_parts:
            if isinstance(current, dict):
                current = current.get(part)
            elif isinstance(current, list) and part.isdigit():
                idx = int(part)
                current = current[idx] if 0 <= idx < len(current) else None
            else:
                return None

        return current


class ExcelImporter(BaseImporter):
    """Excel data importer."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ImportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ImportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process import source and yield batches."""
        if isinstance(source, (str, Path)):
            # Source is a file path
            async for batch in self.import_from_file(Path(source), **kwargs):
                yield batch
        else:
            # Source is a stream
            async for batch in self.import_from_stream(source, **kwargs):
                yield batch

    async def import_from_file(self, file_path: Path, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from Excel file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Select worksheet
            if self.options.sheet_name is not None:
                if isinstance(self.options.sheet_name, str):
                    worksheet = workbook[self.options.sheet_name]
                else:
                    worksheet = workbook.worksheets[self.options.sheet_name]
            else:
                worksheet = workbook.active

            # Process worksheet
            async for batch in self._process_worksheet(worksheet):
                if self._cancelled:
                    break
                yield batch

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ImportError(f"Failed to import Excel file {file_path}: {e}") from e

    async def import_from_stream(
        self, stream: BinaryIO, **kwargs
    ) -> AsyncGenerator[DataBatch, None]:
        """Import data from Excel stream."""
        try:
            # Load workbook from stream
            workbook = openpyxl.load_workbook(stream, data_only=True)

            # Select worksheet
            if self.options.sheet_name is not None:
                if isinstance(self.options.sheet_name, str):
                    worksheet = workbook[self.options.sheet_name]
                else:
                    worksheet = workbook.worksheets[self.options.sheet_name]
            else:
                worksheet = workbook.active

            # Process worksheet
            async for batch in self._process_worksheet(worksheet):
                if self._cancelled:
                    break
                yield batch

        except Exception as e:
            raise ImportError(f"Failed to import from Excel stream: {e}") from e

    async def _process_worksheet(self, worksheet) -> AsyncGenerator[DataBatch, None]:
        """Process Excel worksheet."""
        # Get header row
        header_row = self.options.header_row + 1  # Excel is 1-indexed
        headers = []

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row, column=col)
            header = cell.value or f"Column_{col}"
            headers.append(str(header))

        # Process data rows
        records = []
        for row_idx in range(header_row + 1 + self.options.skip_rows, worksheet.max_row + 1):
            if self._cancelled:
                break

            row_data = {}
            has_data = False

            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                # Convert Excel dates
                if hasattr(cell, "is_date") and cell.is_date and value:
                    # Only call isoformat if it's actually a datetime object
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()

                row_data[header] = value
                if value is not None:
                    has_data = True

            # Skip empty rows
            if not has_data and self.options.skip_blank_lines:
                continue

            record = DataRecord(row_number=row_idx, data=row_data)

            self._validate_record(record)
            record = self._transform_record(record)
            records.append(record)

            # Create batch when size limit reached
            if len(records) >= self.config.batch_size:
                batch = DataBatch(
                    batch_number=self._progress.current_batch + 1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records),
                )

                self._update_progress(
                    processed_records=self._progress.processed_records + len(records),
                    current_batch=self._progress.current_batch + 1,
                )

                yield batch
                records = []

        # Yield remaining records
        if records:
            batch = DataBatch(
                batch_number=self._progress.current_batch + 1,
                records=records,
                total_size=sum(len(str(r.data)) for r in records),
            )

            self._update_progress(
                processed_records=self._progress.processed_records + len(records),
                current_batch=self._progress.current_batch + 1,
            )

            yield batch


class XMLImporter(BaseImporter):
    """XML data importer."""

    def __init__(
        self,
        config: TransferConfig | None = None,
        options: ImportOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ):
        super().__init__(config, progress_callback)
        self.options = options or ImportOptions()

    async def process(self, source: Any, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Process import source and yield batches."""
        if isinstance(source, (str, Path)):
            # Source is a file path
            async for batch in self.import_from_file(Path(source), **kwargs):
                yield batch
        else:
            # Source is a stream
            async for batch in self.import_from_stream(source, **kwargs):
                yield batch

    async def import_from_file(self, file_path: Path, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from XML file."""
        try:
            self._progress.status = TransferStatus.RUNNING

            async with aiofiles.open(file_path, "r", encoding=self.config.encoding) as f:
                content = await f.read()
                async for batch in self._process_xml(content):
                    if self._cancelled:
                        break
                    yield batch

        except Exception as e:
            self._progress.status = TransferStatus.FAILED
            self._progress.error_message = str(e)
            raise ImportError(f"Failed to import XML file {file_path}: {e}") from e

    async def import_from_stream(self, stream: TextIO, **kwargs) -> AsyncGenerator[DataBatch, None]:
        """Import data from XML stream."""
        try:
            # Read content
            if hasattr(stream, "read"):
                # Check if it's an async stream by trying to call read() and checking if it's a coroutine
                try:
                    read_result = stream.read()
                    # If it's a coroutine, await it
                    if hasattr(read_result, "__await__"):
                        content = await read_result
                    else:
                        content = read_result
                except Exception:
                    # Fallback to sync read
                    content = stream.read()
            else:
                content = stream.read()

            async for batch in self._process_xml(content):
                yield batch

        except Exception as e:
            raise ImportError(f"Failed to import from XML stream: {e}") from e

    async def _process_xml(self, content: str) -> AsyncGenerator[DataBatch, None]:
        """Process XML content."""
        try:
            root = ET.fromstring(content)

            # Determine record elements
            if self.options.xml_record_element:
                record_elements = root.findall(f".//{self.options.xml_record_element}")
            else:
                # Use direct children as records
                record_elements = list(root)

            # Process elements
            records = []
            for i, element in enumerate(record_elements):
                if self._cancelled:
                    break

                data = self._element_to_dict(element)
                record = DataRecord(row_number=i + 1, data=data)

                self._validate_record(record)
                record = self._transform_record(record)
                records.append(record)

                # Create batch when size limit reached
                if len(records) >= self.config.batch_size:
                    batch = DataBatch(
                        batch_number=self._progress.current_batch + 1,
                        records=records,
                        total_size=sum(len(str(r.data)) for r in records),
                    )

                    self._update_progress(
                        processed_records=self._progress.processed_records + len(records),
                        current_batch=self._progress.current_batch + 1,
                    )

                    yield batch
                    records = []

            # Yield remaining records
            if records:
                batch = DataBatch(
                    batch_number=self._progress.current_batch + 1,
                    records=records,
                    total_size=sum(len(str(r.data)) for r in records),
                )

                self._update_progress(
                    processed_records=self._progress.processed_records + len(records),
                    current_batch=self._progress.current_batch + 1,
                )

                yield batch

        except ET.ParseError as e:
            raise ImportError(f"Invalid XML format: {e}") from e

    def _element_to_dict(self, element: ET.Element) -> dict[str, Any]:
        """Convert XML element to dictionary."""
        result = {}

        # Add attributes
        if element.attrib:
            result.update({f"@{k}": v for k, v in element.attrib.items()})

        # Add text content
        if element.text and element.text.strip():
            if len(element) == 0:  # No child elements
                return element.text.strip()
            else:
                result["#text"] = element.text.strip()

        # Add child elements
        for child in element:
            child_data = self._element_to_dict(child)

            if child.tag in result:
                # Convert to list if multiple elements with same tag
                if not isinstance(result[child.tag], list):
                    result[child.tag] = [result[child.tag]]
                result[child.tag].append(child_data)
            else:
                result[child.tag] = child_data

        return result


# Factory functions


def create_importer(
    data_format: DataFormat,
    config: TransferConfig | None = None,
    options: ImportOptions | None = None,
    progress_callback: ProgressCallback | None = None,
) -> BaseImporter:
    """Create an appropriate importer for the given format."""
    importers = {
        DataFormat.CSV: CSVImporter,
        DataFormat.JSON: JSONImporter,
        DataFormat.JSONL: lambda config=None, opt=None, progress_cb=None: JSONImporter(
            config=config,
            options=ImportOptions(**(opt.model_dump() if opt else {}), json_lines=True),
            progress_callback=progress_cb,
        ),
        DataFormat.EXCEL: ExcelImporter,
        DataFormat.XML: XMLImporter,
    }

    if data_format not in importers:
        raise FormatError(f"Unsupported format: {data_format}")

    if data_format == DataFormat.JSONL:
        return importers[data_format](config=config, opt=options, progress_cb=progress_callback)
    else:
        return importers[data_format](config, options, progress_callback)


def detect_format(file_path: Path, sample_size: int = 1024) -> DataFormat:
    """Detect file format based on extension and content."""
    # Check file extension first
    suffix = file_path.suffix.lower()
    extension_map = {
        ".csv": DataFormat.CSV,
        ".json": DataFormat.JSON,
        ".jsonl": DataFormat.JSONL,
        ".xlsx": DataFormat.EXCEL,
        ".xls": DataFormat.EXCEL,
        ".xml": DataFormat.XML,
    }

    if suffix in extension_map:
        return extension_map[suffix]

    # Try to detect from content
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            sample = f.read(sample_size)

        # Try JSON
        try:
            json.loads(sample)
            return DataFormat.JSON
        except json.JSONDecodeError:
            pass

        # Try JSON Lines
        lines = sample.strip().split("\n")
        if lines and all(line.strip() for line in lines):
            try:
                for line in lines[:5]:  # Check first 5 lines
                    json.loads(line.strip())
                return DataFormat.JSONL
            except json.JSONDecodeError:
                pass

        # Try XML
        try:
            ET.fromstring(sample)
            return DataFormat.XML
        except ET.ParseError:
            pass

        # Default to CSV
        return DataFormat.CSV

    except Exception:
        # If all else fails, assume CSV
        return DataFormat.CSV


async def import_file(
    file_path: Path,
    data_format: DataFormat | None = None,
    config: TransferConfig | None = None,
    options: ImportOptions | None = None,
    progress_callback: ProgressCallback | None = None,
) -> AsyncGenerator[DataBatch, None]:
    """Import data from file with automatic format detection."""
    if data_format is None:
        data_format = detect_format(file_path)

    importer = create_importer(data_format, config, options, progress_callback)

    async for batch in importer.import_from_file(file_path):
        yield batch
