"""
Comprehensive tests for data transfer streaming and validation functionality.
Targets uncovered streaming methods and validation patterns.
"""

import asyncio
import io
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, AsyncGenerator
from unittest.mock import AsyncMock, Mock, MagicMock, patch

import pytest

from dotmac.platform.data_transfer.base import (
    DataRecord,
    DataBatch,
    DataValidationError,
    StreamingError,
    TransferConfig,
    TransferStatus,
    ProgressInfo,
    BaseImporter,
    BaseExporter,
)
from dotmac.platform.data_transfer.exporters import (
    CSVExporter,
    JSONExporter,
    XMLExporter,
    YAMLExporter,
    ExcelExporter,
    ExportOptions,
)
from dotmac.platform.data_transfer.importers import (
    CSVImporter,
    JSONImporter,
    XMLImporter,
    ExcelImporter,
    ImportOptions,
)


class TestDataRecordValidation:
    """Test DataRecord validation functionality."""

    def test_data_record_validation_errors_basic(self):
        """Test basic validation errors on data record."""
        record = DataRecord(
            id="test1",
            data={"name": "John"},
            validation_errors=["Email required", "Phone invalid"],
            is_valid=False
        )

        assert record.validation_errors == ["Email required", "Phone invalid"]
        assert record.is_valid is False

    def test_data_record_validation_errors_empty(self):
        """Test data record with no validation errors."""
        record = DataRecord(
            id="test2",
            data={"name": "John", "email": "john@test.com"},
            validation_errors=[],
            is_valid=True
        )

        assert record.validation_errors == []
        assert record.is_valid is True

    def test_data_record_validation_errors_modification(self):
        """Test modification of validation errors."""
        record = DataRecord(
            id="test3",
            data={"name": "John"},
            validation_errors=["Initial error"],
            is_valid=True
        )

        # Add more validation errors
        record.validation_errors.append("Second error")
        record.validation_errors.append("Third error")
        record.is_valid = False

        assert len(record.validation_errors) == 3
        assert "Initial error" in record.validation_errors
        assert "Second error" in record.validation_errors
        assert "Third error" in record.validation_errors
        assert record.is_valid is False

    def test_data_record_with_row_number(self):
        """Test data record with row number for tracking."""
        record = DataRecord(
            id="test4",
            data={"field1": "value1"},
            row_number=42,
            validation_errors=["Parse error"],
            is_valid=False
        )

        assert record.row_number == 42
        assert record.validation_errors == ["Parse error"]

    def test_data_record_empty_data_validation(self):
        """Test validation of empty data."""
        with pytest.raises(ValueError, match="Record data cannot be empty"):
            DataRecord(id="test5", data={})


class TestDataBatchValidation:
    """Test DataBatch validation functionality."""

    def test_data_batch_valid_records_property(self):
        """Test valid_records property filtering."""
        records = [
            DataRecord(id="1", data={"field": "value1"}, is_valid=True),
            DataRecord(id="2", data={"field": "value2"}, is_valid=False),
            DataRecord(id="3", data={"field": "value3"}, is_valid=True),
            DataRecord(id="4", data={"field": "value4"}, is_valid=False),
        ]

        batch = DataBatch(records=records)
        valid_records = batch.valid_records

        assert len(valid_records) == 2
        assert valid_records[0].id == "1"
        assert valid_records[1].id == "3"
        assert all(record.is_valid for record in valid_records)

    def test_data_batch_invalid_records_property(self):
        """Test invalid_records property filtering."""
        records = [
            DataRecord(id="1", data={"field": "value1"}, is_valid=True),
            DataRecord(id="2", data={"field": "value2"}, is_valid=False),
            DataRecord(id="3", data={"field": "value3"}, is_valid=True),
            DataRecord(id="4", data={"field": "value4"}, is_valid=False),
        ]

        batch = DataBatch(records=records)
        invalid_records = batch.invalid_records

        assert len(invalid_records) == 2
        assert invalid_records[0].id == "2"
        assert invalid_records[1].id == "4"
        assert all(not record.is_valid for record in invalid_records)

    def test_data_batch_all_valid_records(self):
        """Test batch with all valid records."""
        records = [
            DataRecord(id="1", data={"field": "value1"}, is_valid=True),
            DataRecord(id="2", data={"field": "value2"}, is_valid=True),
        ]

        batch = DataBatch(records=records)

        assert len(batch.valid_records) == 2
        assert len(batch.invalid_records) == 0

    def test_data_batch_all_invalid_records(self):
        """Test batch with all invalid records."""
        records = [
            DataRecord(id="1", data={"field": "value1"}, is_valid=False),
            DataRecord(id="2", data={"field": "value2"}, is_valid=False),
        ]

        batch = DataBatch(records=records)

        assert len(batch.valid_records) == 0
        assert len(batch.invalid_records) == 2

    def test_data_batch_empty_records(self):
        """Test batch with no records."""
        batch = DataBatch(records=[])

        assert len(batch.valid_records) == 0
        assert len(batch.invalid_records) == 0


class TestBaseExporterValidation:
    """Test BaseExporter validation functionality."""

    class MockExporter(BaseExporter):
        """Mock exporter for testing validation."""

        def __init__(self, config: TransferConfig = None, **kwargs):
            super().__init__(config, **kwargs)
            self.validator = kwargs.get("validator")
            self.transformer = kwargs.get("transformer")

        async def export_to_file(self, data: AsyncGenerator[DataBatch, None], file_path: Path, **kwargs) -> ProgressInfo:
            """Mock export to file."""
            return self._progress

        async def export_to_stream(self, data: AsyncGenerator[DataBatch, None], stream: Any, **kwargs) -> ProgressInfo:
            """Mock export to stream."""
            return self._progress

    @pytest.fixture
    def mock_config(self):
        """Create mock transfer config."""
        return settings.Transfer.model_copy(update={validate_data=True, batch_size=10})

    @pytest.fixture
    def mock_exporter(self, mock_config):
        """Create mock exporter instance."""
        return self.MockExporter(mock_config)

    def test_validate_record_with_validation_disabled(self, mock_config):
        """Test record validation when validation is disabled."""
        mock_config.validate_data = False
        exporter = self.MockExporter(mock_config)

        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._validate_record(record)

        assert result is True
        assert record.is_valid is True

    def test_validate_record_with_no_validator(self, mock_exporter):
        """Test record validation when no validator is set."""
        record = DataRecord(id="test", data={"field": "value"})
        result = mock_exporter._validate_record(record)

        assert result is True
        assert record.is_valid is True

    def test_validate_record_with_successful_validator(self, mock_config):
        """Test record validation with successful validator."""
        def success_validator(record):
            return True

        exporter = self.MockExporter(mock_config, validator=success_validator)
        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._validate_record(record)

        assert result is True
        assert record.is_valid is True
        assert len(record.validation_errors) == 0

    def test_validate_record_with_failing_validator(self, mock_config):
        """Test record validation with failing validator."""
        def fail_validator(record):
            return False

        exporter = self.MockExporter(mock_config, validator=fail_validator)
        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._validate_record(record)

        assert result is False
        assert record.is_valid is False

    def test_validate_record_with_exception_in_validator(self, mock_config):
        """Test record validation when validator raises exception."""
        def exception_validator(record):
            raise ValueError("Validation failed with custom error")

        exporter = self.MockExporter(mock_config, validator=exception_validator)
        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._validate_record(record)

        assert result is False
        assert record.is_valid is False
        assert "Validation failed with custom error" in record.validation_errors

    def test_transform_record_with_no_transformer(self, mock_exporter):
        """Test record transformation when no transformer is set."""
        record = DataRecord(id="test", data={"field": "original"})
        result = mock_exporter._transform_record(record)

        assert result == record
        assert result.data["field"] == "original"

    def test_transform_record_with_successful_transformer(self, mock_config):
        """Test record transformation with successful transformer."""
        def success_transformer(record):
            record.data["transformed"] = True
            return record

        exporter = self.MockExporter(mock_config, transformer=success_transformer)
        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._transform_record(record)

        assert result == record
        assert record.data["transformed"] is True
        assert record.is_valid is True

    def test_transform_record_with_exception_in_transformer(self, mock_config):
        """Test record transformation when transformer raises exception."""
        def exception_transformer(record):
            raise RuntimeError("Transformation failed")

        exporter = self.MockExporter(mock_config, transformer=exception_transformer)
        record = DataRecord(id="test", data={"field": "value"})
        result = exporter._transform_record(record)

        assert result == record
        assert record.is_valid is False
        assert "Transformation error: Transformation failed" in record.validation_errors


class TestStreamingExporters:
    """Test streaming functionality in exporters."""

    @pytest.fixture
    def sample_data_generator(self):
        """Create sample data generator for testing."""
        async def generator():
            batch1 = DataBatch(records=[
                DataRecord(id="1", data={"name": "John", "age": 30}),
                DataRecord(id="2", data={"name": "Jane", "age": 25}),
            ])
            yield batch1

            batch2 = DataBatch(records=[
                DataRecord(id="3", data={"name": "Bob", "age": 35}),
            ])
            yield batch2

        return generator()

    @pytest.mark.asyncio
    async def test_csv_exporter_stream_basic(self, sample_data_generator):
        """Test CSV exporter streaming functionality."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions(include_headers=True, delimiter=",")
        exporter = CSVExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(sample_data_generator, stream)

        # Verify the result
        assert progress.status == TransferStatus.COMPLETED
        assert progress.processed_records == 3

        # Check stream content
        content = stream.getvalue()
        lines = content.strip().split('\n')
        assert len(lines) == 4  # Header + 3 data rows
        assert "name,age" in lines[0]  # Header
        assert "John,30" in content
        assert "Jane,25" in content
        assert "Bob,35" in content

    @pytest.mark.asyncio
    async def test_json_exporter_stream_standard_format(self, sample_data_generator):
        """Test JSON exporter streaming with standard JSON format."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions(json_lines=False, json_indent=2)
        exporter = JSONExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(sample_data_generator, stream)

        # Verify the result
        assert progress.status == TransferStatus.COMPLETED
        assert progress.processed_records == 3

        # Check stream content
        content = stream.getvalue()
        data = json.loads(content)
        assert isinstance(data, list)
        assert len(data) == 3
        assert data[0]["name"] == "John"
        assert data[0]["age"] == 30

    @pytest.mark.asyncio
    async def test_json_exporter_stream_jsonl_format(self, sample_data_generator):
        """Test JSON exporter streaming with JSON Lines format."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions(json_lines=True)
        exporter = JSONExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(sample_data_generator, stream)

        # Verify the result
        assert progress.status == TransferStatus.COMPLETED
        assert progress.processed_records == 3

        # Check stream content
        content = stream.getvalue()
        lines = [line.strip() for line in content.strip().split('\n') if line.strip()]
        assert len(lines) == 3

        # Parse each line as JSON
        for line in lines:
            data = json.loads(line)
            assert "name" in data
            assert "age" in data

    @pytest.mark.asyncio
    async def test_xml_exporter_stream_basic(self, sample_data_generator):
        """Test XML exporter streaming functionality."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions(xml_root_element="people")
        exporter = XMLExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(sample_data_generator, stream)

        # Verify the result
        assert progress.status == TransferStatus.COMPLETED
        assert progress.processed_records == 3

        # Check stream content
        content = stream.getvalue()
        assert '<?xml version="1.0"' in content
        assert "<people>" in content
        assert "</people>" in content
        assert "<name>John</name>" in content
        assert "<age>30</age>" in content

    @pytest.mark.asyncio
    async def test_yaml_exporter_stream_basic(self, sample_data_generator):
        """Test YAML exporter streaming functionality."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions()
        exporter = YAMLExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(sample_data_generator, stream)

        # Verify the result
        assert progress.status == TransferStatus.COMPLETED
        assert progress.processed_records == 3

        # Check stream content
        content = stream.getvalue()
        assert "name: John" in content
        assert "age: 30" in content
        assert "name: Jane" in content
        assert "age: 25" in content

    @pytest.mark.asyncio
    async def test_excel_exporter_stream_not_supported(self, sample_data_generator):
        """Test that Excel exporter streaming raises appropriate error."""
        config = settings.Transfer.model_copy(update={validate_data=False})
        options = ExportOptions(sheet_name="TestSheet")
        exporter = ExcelExporter(config, options)

        stream = io.BytesIO()

        with pytest.raises(Exception, match="Excel export to stream not supported"):
            await exporter.export_to_stream(sample_data_generator, stream)

    @pytest.mark.asyncio
    async def test_stream_export_with_validation_enabled(self):
        """Test streaming export with validation enabled."""
        async def data_with_invalid():
            batch = DataBatch(records=[
                DataRecord(id="1", data={"name": "John", "age": 30}, is_valid=True),
                DataRecord(id="2", data={"name": "", "age": -5}, is_valid=False),  # Invalid
                DataRecord(id="3", data={"name": "Jane", "age": 25}, is_valid=True),
            ])
            yield batch

        config = settings.Transfer.model_copy(update={validate_data=True, skip_invalid=True})
        options = ExportOptions(include_headers=True)
        exporter = CSVExporter(config, options)

        stream = io.StringIO()
        progress = await exporter.export_to_stream(data_with_invalid(), stream)

        # Should skip invalid record
        assert progress.processed_records == 2  # Only valid records processed

        content = stream.getvalue()
        assert "John" in content
        assert "Jane" in content
        assert '"","-5"' not in content  # Invalid record not included

    @pytest.mark.asyncio
    async def test_stream_export_cancellation(self):
        """Test streaming export cancellation."""
        async def large_data_generator():
            for i in range(1000):  # Large dataset
                batch = DataBatch(records=[
                    DataRecord(id=str(i), data={"index": i, "data": f"value_{i}"})
                ])
                yield batch

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=10})
        exporter = CSVExporter(config)

        # Cancel after starting
        exporter.cancel()

        stream = io.StringIO()
        progress = await exporter.export_to_stream(large_data_generator(), stream)

        # Should have stopped early due to cancellation
        assert progress.processed_records < 1000


class TestStreamingImporters:
    """Test streaming functionality in importers."""

    @pytest.mark.asyncio
    async def test_csv_importer_stream_basic(self):
        """Test CSV importer streaming functionality."""
        csv_content = "name,age,city\nJohn,30,NYC\nJane,25,LA\nBob,35,Chicago"
        stream = io.StringIO(csv_content)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=2})
        options = ImportOptions(has_header=True, delimiter=",")
        importer = CSVImporter(config, options)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 2 batches (2 records each, except last)
        assert len(batches) == 2
        assert len(batches[0].records) == 2
        assert len(batches[1].records) == 1

        # Check data
        first_record = batches[0].records[0]
        assert first_record.data["name"] == "John"
        assert first_record.data["age"] == "30"
        assert first_record.data["city"] == "NYC"

    @pytest.mark.asyncio
    async def test_json_importer_stream_array_format(self):
        """Test JSON importer streaming with array format."""
        json_content = '[{"name": "John", "age": 30}, {"name": "Jane", "age": 25}]'
        stream = io.StringIO(json_content)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=1})
        importer = JSONImporter(config)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 2 batches (1 record each)
        assert len(batches) == 2
        assert batches[0].records[0].data["name"] == "John"
        assert batches[1].records[0].data["name"] == "Jane"

    @pytest.mark.asyncio
    async def test_json_importer_stream_jsonl_format(self):
        """Test JSON importer streaming with JSON Lines format."""
        jsonl_content = '{"name": "John", "age": 30}\n{"name": "Jane", "age": 25}\n'
        stream = io.StringIO(jsonl_content)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=1})
        importer = JSONImporter(config)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 2 batches (1 record each)
        assert len(batches) == 2
        assert batches[0].records[0].data["name"] == "John"
        assert batches[1].records[0].data["name"] == "Jane"

    @pytest.mark.asyncio
    async def test_json_importer_stream_invalid_jsonl(self):
        """Test JSON importer with invalid JSON Lines."""
        invalid_jsonl = '{"name": "John", "age": 30}\n{"invalid": json}\n{"name": "Jane", "age": 25}\n'
        stream = io.StringIO(invalid_jsonl)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=10})
        importer = JSONImporter(config)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 1 batch with 3 records (including error record)
        assert len(batches) == 1
        batch = batches[0]
        assert len(batch.records) == 3

        # Check valid records
        assert batch.records[0].is_valid is True
        assert batch.records[0].data["name"] == "John"

        # Check invalid record
        assert batch.records[1].is_valid is False
        assert len(batch.records[1].validation_errors) > 0
        assert "JSON decode error" in batch.records[1].validation_errors[0]

        # Check last valid record
        assert batch.records[2].is_valid is True
        assert batch.records[2].data["name"] == "Jane"

    @pytest.mark.asyncio
    async def test_excel_importer_stream_basic(self):
        """Test Excel importer streaming functionality."""
        # Create mock Excel content in memory
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Add headers
        ws['A1'] = 'name'
        ws['B1'] = 'age'
        ws['C1'] = 'city'

        # Add data
        ws['A2'] = 'John'
        ws['B2'] = 30
        ws['C2'] = 'NYC'

        ws['A3'] = 'Jane'
        ws['B3'] = 25
        ws['C3'] = 'LA'

        # Save to stream
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=10})
        options = ImportOptions(sheet_name="TestSheet", has_header=True)
        importer = ExcelImporter(config, options)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 1 batch with 2 records
        assert len(batches) == 1
        batch = batches[0]
        assert len(batch.records) == 2

        # Check data
        assert batch.records[0].data["name"] == "John"
        assert batch.records[0].data["age"] == 30
        assert batch.records[1].data["name"] == "Jane"
        assert batch.records[1].data["age"] == 25

    @pytest.mark.asyncio
    async def test_xml_importer_stream_basic(self):
        """Test XML importer streaming functionality."""
        xml_content = '''<?xml version="1.0" encoding="utf-8"?>
        <records>
            <record>
                <name>John</name>
                <age>30</age>
            </record>
            <record>
                <name>Jane</name>
                <age>25</age>
            </record>
        </records>'''

        stream = io.StringIO(xml_content)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=10})
        options = ImportOptions(xml_root_element="records", xml_record_element="record")
        importer = XMLImporter(config, options)

        batches = []
        async for batch in importer.import_from_stream(stream):
            batches.append(batch)

        # Should have 1 batch with 2 records
        assert len(batches) == 1
        batch = batches[0]
        assert len(batch.records) == 2

        # Check data
        assert batch.records[0].data["name"] == "John"
        assert batch.records[0].data["age"] == "30"
        assert batch.records[1].data["name"] == "Jane"
        assert batch.records[1].data["age"] == "25"


class TestStreamingErrorHandling:
    """Test error handling in streaming operations."""

    @pytest.mark.asyncio
    async def test_streaming_error_exception(self):
        """Test StreamingError exception handling."""
        with pytest.raises(StreamingError):
            raise StreamingError("Test streaming error")

    @pytest.mark.asyncio
    async def test_data_validation_error_exception(self):
        """Test DataValidationError exception handling."""
        with pytest.raises(DataValidationError):
            raise DataValidationError("Test validation error")

    @pytest.mark.asyncio
    async def test_exporter_stream_write_error(self):
        """Test error handling during stream write."""
        class FailingStream:
            def write(self, content):
                raise IOError("Stream write failed")

        async def sample_data():
            batch = DataBatch(records=[
                DataRecord(id="1", data={"name": "John", "age": 30})
            ])
            yield batch

        config = settings.Transfer.model_copy(update={validate_data=False})
        exporter = CSVExporter(config)

        failing_stream = FailingStream()

        with pytest.raises(Exception):  # Should propagate the stream error
            await exporter.export_to_stream(sample_data(), failing_stream)

    @pytest.mark.asyncio
    async def test_importer_stream_read_error(self):
        """Test error handling during stream read."""
        class FailingStream:
            def read(self):
                raise IOError("Stream read failed")

        config = settings.Transfer.model_copy(update={validate_data=False})
        importer = CSVImporter(config)

        failing_stream = FailingStream()

        with pytest.raises(Exception):  # Should propagate the stream error
            async for batch in importer.import_from_stream(failing_stream):
                pass

    @pytest.mark.asyncio
    async def test_async_stream_handling(self):
        """Test handling of async streams."""
        class AsyncStream:
            def __init__(self, content):
                self.content = content
                self.position = 0

            async def read(self):
                if self.position >= len(self.content):
                    return ""
                result = self.content[self.position:]
                self.position = len(self.content)
                return result

        csv_content = "name,age\nJohn,30\nJane,25"
        async_stream = AsyncStream(csv_content)

        config = settings.Transfer.model_copy(update={validate_data=False, batch_size=10})
        options = ImportOptions(has_header=True)
        importer = CSVImporter(config, options)

        batches = []
        async for batch in importer.import_from_stream(async_stream):
            batches.append(batch)

        # Should successfully process async stream
        assert len(batches) == 1
        assert len(batches[0].records) == 2


class TestValidationEdgeCases:
    """Test validation edge cases and error scenarios."""

    def test_validation_config_disabled(self):
        """Test validation behavior when disabled in config."""
        config = settings.Transfer.model_copy(update={validate_data=False})

        # Even invalid data should pass when validation is disabled
        record = DataRecord(id="test", data={"field": "value"})

        class TestExporter(BaseExporter):
            async def export_to_file(self, data, file_path, **kwargs):
                return self._progress
            async def export_to_stream(self, data, stream, **kwargs):
                return self._progress

        exporter = TestExporter(config)
        result = exporter._validate_record(record)

        assert result is True
        assert record.is_valid is True

    def test_custom_validator_complex_logic(self):
        """Test custom validator with complex validation logic."""
        def complex_validator(record):
            data = record.data

            # Multiple validation rules
            if not data.get("email") or "@" not in data.get("email", ""):
                record.validation_errors.append("Invalid email format")
                record.is_valid = False
                return False

            if not data.get("age") or int(data.get("age", 0)) < 18:
                record.validation_errors.append("Age must be 18 or older")
                record.is_valid = False
                return False

            return True

        config = settings.Transfer.model_copy(update={validate_data=True})

        class TestExporter(BaseExporter):
            async def export_to_file(self, data, file_path, **kwargs):
                return self._progress
            async def export_to_stream(self, data, stream, **kwargs):
                return self._progress

        exporter = TestExporter(config, validator=complex_validator)

        # Test valid record
        valid_record = DataRecord(id="1", data={"email": "test@example.com", "age": "25"})
        result = exporter._validate_record(valid_record)
        assert result is True
        assert valid_record.is_valid is True

        # Test invalid email
        invalid_email_record = DataRecord(id="2", data={"email": "invalid", "age": "25"})
        result = exporter._validate_record(invalid_email_record)
        assert result is False
        assert "Invalid email format" in invalid_email_record.validation_errors

        # Test invalid age
        invalid_age_record = DataRecord(id="3", data={"email": "test@example.com", "age": "15"})
        result = exporter._validate_record(invalid_age_record)
        assert result is False
        assert "Age must be 18 or older" in invalid_age_record.validation_errors

    def test_transformer_with_validation_interaction(self):
        """Test interaction between transformer and validation."""
        def email_normalizer(record):
            if "email" in record.data:
                record.data["email"] = record.data["email"].lower().strip()
            return record

        def email_validator(record):
            email = record.data.get("email", "")
            if not email or "@" not in email:
                record.validation_errors.append("Invalid email")
                record.is_valid = False
                return False
            return True

        config = settings.Transfer.model_copy(update={validate_data=True})

        class TestExporter(BaseExporter):
            async def export_to_file(self, data, file_path, **kwargs):
                return self._progress
            async def export_to_stream(self, data, stream, **kwargs):
                return self._progress

        exporter = TestExporter(config, transformer=email_normalizer, validator=email_validator)

        # Test record that needs transformation
        record = DataRecord(id="test", data={"email": "  TEST@EXAMPLE.COM  "})

        # Transform first
        transformed = exporter._transform_record(record)
        assert transformed.data["email"] == "test@example.com"

        # Then validate
        result = exporter._validate_record(transformed)
        assert result is True
        assert transformed.is_valid is True